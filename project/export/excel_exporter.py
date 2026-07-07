"""분석 결과를 Excel 파일로 저장한다.

Sheet1(Result): 원본 데이터 + 이상치 여부 + 이상치 사유, 이상 행은 연한 노란색.
Sheet2(Anomaly): 이상 데이터만 추출.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from ..config import Settings
from ..models import AnalysisResult


class ExcelExporter:
    """AnalysisResult 를 스타일이 적용된 Excel 로 내보내는 클래스."""

    def __init__(self, settings: Settings | None = None) -> None:
        self._settings = settings or Settings()

    def export(self, result: AnalysisResult, output_path: str) -> str:
        """결과를 Excel 파일로 저장한다.

        Args:
            result: 분석 결과.
            output_path: 저장할 파일 경로(.xlsx).

        Returns:
            실제 저장된 파일 경로.
        """
        if not output_path.lower().endswith(".xlsx"):
            output_path = f"{output_path}.xlsx"

        # 결과 컬럼을 추가한 DataFrame 구성(원본은 유지).
        result_df = result.data.copy()
        result_df[self._settings.ANOMALY_FLAG_COLUMN] = result.anomaly_flags
        result_df[self._settings.ANOMALY_REASON_COLUMN] = result.anomaly_reasons

        workbook = Workbook()
        result_sheet = workbook.active
        result_sheet.title = self._settings.RESULT_SHEET_NAME

        self._write_sheet(result_sheet, result_df, highlight=True)

        # Sheet2: 이상 데이터만
        anomaly_df = result_df[
            result_df[self._settings.ANOMALY_FLAG_COLUMN]
            == self._settings.ANOMALY_FLAG_VALUE
        ].reset_index(drop=True)
        anomaly_sheet = workbook.create_sheet(title=self._settings.ANOMALY_SHEET_NAME)
        self._write_sheet(anomaly_sheet, anomaly_df, highlight=True)

        workbook.save(output_path)
        return output_path

    def _write_sheet(
        self, sheet: Worksheet, df: pd.DataFrame, highlight: bool
    ) -> None:
        """DataFrame 을 시트에 기록하고 스타일을 적용한다."""
        fill = PatternFill(
            start_color=self._settings.ANOMALY_HIGHLIGHT_COLOR,
            end_color=self._settings.ANOMALY_HIGHLIGHT_COLOR,
            fill_type="solid",
        )
        header_font = Font(bold=True)

        flag_col_name = self._settings.ANOMALY_FLAG_COLUMN
        flag_value = self._settings.ANOMALY_FLAG_VALUE
        flag_col_pos = (
            list(df.columns).index(flag_col_name) if flag_col_name in df.columns else -1
        )

        # 헤더 + 데이터 기록
        for row_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=True), start=1
        ):
            is_anomaly_row = False
            if highlight and flag_col_pos >= 0 and row_idx > 1:
                is_anomaly_row = row[flag_col_pos] == flag_value

            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=self._to_cell(value))
                if row_idx == 1:
                    cell.font = header_font
                elif is_anomaly_row:
                    cell.fill = fill

        self._auto_size_columns(sheet, df)

    @staticmethod
    def _to_cell(value: object) -> object:
        """openpyxl 이 처리할 수 있는 값으로 정규화한다."""
        if pd.isna(value):
            return None
        # numpy 스칼라 등은 파이썬 기본형으로 변환
        if hasattr(value, "item"):
            try:
                return value.item()
            except (ValueError, AttributeError):
                return value
        return value

    @staticmethod
    def _auto_size_columns(sheet: Worksheet, df: pd.DataFrame) -> None:
        """컬럼 너비를 내용 길이에 맞춰 대략 조정한다."""
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name))
            # 성능을 위해 상위 일부만 샘플링
            sample = df[col_name].head(200)
            for value in sample:
                if pd.isna(value):
                    continue
                max_len = max(max_len, len(str(value)))
            width = min(max_len + 2, 60)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
